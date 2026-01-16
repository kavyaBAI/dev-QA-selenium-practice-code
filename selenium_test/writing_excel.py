import openpyxl

file=r"/home/kavr/Desktop/QA-AUTOMATION/DEV_test/selenium_test/data_1.xlsx"
workbook=openpyxl.load_workbook(file)


##############################writing single vale to excel###########################################
#sheet=workbook["data2"]
# for r in range(1,6):
#     for c in range(1,5):
#         sheet.cell(r,c).value="welcome"
# workbook.save(file)
# print("Data written successfully.")

###############################################################################################

#multiple data
sheet=workbook["data3"]
#Writing first row of data
sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="engineer"

# Writing second row of data
sheet.cell(2,1).value=567
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

# Writing third row of data
sheet.cell(3,1).value=567
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="manager" 

workbook.save(file)
print("Data written successfully 2.")
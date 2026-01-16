import openpyxl

#File-->Workbook-->Sheets-->Rows-->Cells

file=r"/home/kavr/Desktop/QA-AUTOMATION/DEV_test/selenium_test/data_1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

rows=sheet.max_row # count number of rows in a excel sheet
cols=sheet.max_column # count number of columns in a excel sheet

#Reading all the rows & columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='   ')
    print()